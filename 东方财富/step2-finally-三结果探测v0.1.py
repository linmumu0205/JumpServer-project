#import os
import openpyxl
import paramiko
import time
from winrm.protocol import Protocol
from concurrent.futures import ThreadPoolExecutor
import logging

# 配置日志记录器
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 全局锁，用于确保对Excel文件的写操作是线程安全的
import threading

excel_lock = threading.Lock()


def ssh_probe(hostip, username='your_username', password='your_password', sshport=1):
    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(hostname=hostip, port=sshport, username=username, password=password, timeout=20)
        stdin, stdout, stderr = ssh.exec_command('echo 1')  # 轻量级探测命令
        output = stdout.read().decode('utf-8')
        ssh.close()
        return True, output
    except paramiko.AuthenticationException:
        return False, "Authentication failed."
    except paramiko.SSHException as e:
        return False, f"SSH connection failed: {e}"
    except Exception as e:
        return False, f"Unexpected error: {e}"


def check_rdp_login(hostip, username='your_username', password='your_password'):
    # 创建 WinRM 连接对象
    winrm = Protocol(
        endpoint=f'http://{hostip}:5985/wsman',
        transport='ntlm',
        username=username,
        password=password,
        server_cert_validation='ignore',
        operation_timeout_sec=20
    )

    shell_id = None  # 赋予初始值；避免try失败情况下，导致shell_id未被赋值
    try:
        # 尝试建立连接
        shell_id = winrm.open_shell()
        return True
    except Exception as e:
        logger.error(f"登录失败：{e}")
        return False
    finally:
        if shell_id:
            winrm.close_shell(shell_id)


def test_connectivity(hostname, hostip, username, password, protocol, ws, idx, row, sshport):
    result_column_index_ssh45685 = 12
    result_column_index_ssh22 = 13
    result_column_index_rdp = 14
    success, result = ssh_probe(hostip, username, password, sshport=45685)
    if success:
        logger.info(f"SSH probe successful for {idx}:{hostip}{hostname}: {result}")
        with excel_lock:
            ws.cell(row=idx, column=result_column_index_ssh45685, value="ssh_success")
    else:
        logger.info(f"SSH probe failed for {idx}:{hostip}{hostname}: {result}")
        with excel_lock:
            ws.cell(row=idx, column=result_column_index_ssh45685, value=f"failed: {result}")

    success, result = ssh_probe(hostip, username, password, sshport=22)
    if success:
        logger.info(f"SSH probe successful for {idx}:{hostip}{hostname}: {result}")
        with excel_lock:
            ws.cell(row=idx, column=result_column_index_ssh22, value="ssh_success")
    else:
        logger.info(f"SSH probe failed for {idx}:{hostip}{hostname}: {result}")
        with excel_lock:
            ws.cell(row=idx, column=result_column_index_ssh22, value=f"failed: {result}")
            
    if check_rdp_login(hostip, username, password):
        logger.info(f"{idx}: 可以使用账号 {username} 登录到主机 {hostip}。")
        with excel_lock:
            ws.cell(row=idx, column=result_column_index_rdp, value="rdp_success")
    else:
        logger.info(f"{idx}: 无法使用账号 {username} 登录到主机 {hostip}。")
        with excel_lock:
            ws.cell(row=idx, column=result_column_index_rdp, value="rdp_failed")


def test_sshrdp_connectivity(file_path, new_file_path):
    start_time = time.time()
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 默认使用10个max_workers,调整workers控制并发线程
    with ThreadPoolExecutor(max_workers=500) as executor:
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            hostname, hostip, username, password, protocol, sshport = row[1], row[2], row[6], row[7], row[3], row[3]
            executor.submit(test_connectivity, hostname, hostip, username, password, protocol, ws, idx, row, sshport)
            # 更新Excel表格的结果

    wb.save(new_file_path)
    end_time = time.time()
    logger.info(f"所有主机探测完成，总耗时: {end_time - start_time:.2f}秒")


if __name__ == "__main__":
    #script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = "/opt/jumpserver/JumpServer-Asset-2024-04-11.xlsx"
    new_file_path = "/opt/jumpserver/JumpServer-Asset-2024-04-11-result002.xlsx"
    test_sshrdp_connectivity(file_path, new_file_path)

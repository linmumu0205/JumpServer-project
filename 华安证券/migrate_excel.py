# -*- coding: utf-8 -*-
#
from collections import OrderedDict

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

excel_table = 'JumpServer-Asset-2024-05-13.xlsx'
sheet_name = 'Sheet'


class Tool(object):
    def __init__(self, excel, sheet):
        self.correct_data = []
        self.correct_assets = set()
        self.incorrect_assets = set()
        self.sheet_name = sheet
        self.wb = load_workbook(excel, read_only=True)
        self.ws = self.wb[self.sheet_name]
        self.max_row = self.ws.max_row

    def read_raw_excel(self, rs):
        if str(rs[1]).strip() == str(rs[5]).strip():
            print("对应成功：%s %s\n" % (rs[0], str(rs[1]).strip()))
            self.correct_assets.add((rs[0], str(rs[1]).strip(), str(rs[2]).strip()))
            self.correct_data.append(rs)
        else:
            self.incorrect_assets.add((rs[0], str(rs[1]).strip(), str(rs[2]).strip()))

    @staticmethod
    def __auto_adjust_column_width(ws):
        def calc_handler(x):
            return len(x.encode()) if x else 0

        pre_rows = ws.iter_cols(min_row=1, max_row=2, values_only=True)
        # 动态分配表头宽度
        column_width = [max(map(calc_handler, i)) for i in pre_rows]
        # 调整首行宽度
        for i, width in enumerate(column_width, 1):
            width = width if width < 100 else 100
            width = 10 if width < 10 else width
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = width

    def write_correct_result(self):
        header = OrderedDict({
            'id': '主机ID', 'hostname': '主机名', 'ip': 'IP', 'asset_protocol': '资产协议', 'system_user_id': '账号ID',
            'system_user': '系统用户名称', 'system_user_protocol': '系统用户协议',
            'username': '账号用户名', 'password': '密码', 'private_key': '秘钥', 'public_key': '公钥',
            'login_date': '最近一次登录时间', 'org_id': '组织ID', 'org_name': '组织名称'
        })
        filename = f'./JumpServer-Correct-Asset.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.freeze_panes = 'A2'
        ws.append(list(header.values()))
        for d in self.correct_data:
            if len(d) < 1:
                continue
            ws.append(d)
        self.__auto_adjust_column_width(ws)
        wb.save(filename)

    def write_incorrect_result(self):
        filename = f'./JumpServer-Incorrect-Asset.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.freeze_panes = 'A2'
        ws.append(['主机ID', '主机名', 'IP'])
        for d in self.incorrect_assets - self.correct_assets:
            if len(d) < 1:
                continue
            ws.append(d)
        self.__auto_adjust_column_width(ws)
        wb.save(filename)

    def run(self):
        print("-" * 20 + "开始处理" + "-" * 20)
        rs = self.ws.iter_rows(values_only=True, max_row=self.max_row, min_row=2)
        from concurrent.futures import ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=20) as executor:
            executor.map(self.read_raw_excel, rs)
        print("-" * 20 + "处理已完成" + "-" * 20)
        print("正确对应资产账号数量：%d" % len(self.correct_data))
        print("正确资产数量：%d" % len(self.correct_assets))
        print("不正确资产数量：%d" % len(self.incorrect_assets - self.correct_assets))
        self.write_correct_result()
        print("写入对应成功数据成功 ./JumpServer-Correct-Asset.xlsx")
        self.write_incorrect_result()
        print("写入对应成功数据成功 ./JumpServer-InCorrect-Asset.xlsx")


if __name__ == '__main__':
    try:
        t = Tool(excel_table, sheet_name)
        t.run()
    except KeyboardInterrupt:
        print('用户终止程序')

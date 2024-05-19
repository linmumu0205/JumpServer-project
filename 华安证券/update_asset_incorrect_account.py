# -*- coding: utf-8 -*-
#

import uuid

import requests
from openpyxl import load_workbook

server_url = 'http://localhost:8080'
token = 'b5667101af57d4ebc14c7622986e1d199609c7ad'
org_id = '00000000-0000-0000-0000-000000000002'
correct_excel = './JumpServer-Incorrect-Asset.xlsx'


class Tool(object):
    def __init__(self):
        self.is_end = False
        self.wb = load_workbook(correct_excel)

    def get_row_value(self, sheet_name, raw_no):
        """
        获取某一行的数据
        sheet_name是excel最下面的Sheet1、Sheet2、Sheet3.....
        """
        sh = self.wb[sheet_name]
        row_value_list = []
        for y in range(1, sh.max_column + 1):
            value = sh.cell(raw_no, y).value
            # 检测读到了文件尾
            if y == 2 and value is None:
                self.is_end = value is None
                break
            row_value_list.append(value)
        return row_value_list

    def get_data_from_excel(self):
        data = []
        for y in range(2, 99999):
            row_value_list = self.get_row_value('Sheet', y)
            if self.is_end:
                break
            else:
                data.append(row_value_list)
        return data

    def create_account_passwd(self, ac):
        url = server_url + '/api/v1/accounts/accounts/'
        headers = {
            'content-type': 'application/json',
            'X-JMS-ORG': org_id,
            'Authorization': 'Token ' + token  # 认证 token
        }
        playload = {
            'asset': str(uuid.UUID(ac['asset_id'])),
            'username': ac['username'],
            'name': ac['name'],
            'privileged': ac['privileged'],
            'secret_type': 'password',
            'on_invalid': 'error',
            'is_active': True,
            'secret': ac['secret']
        }
        response = requests.request('POST', url, json=playload, headers=headers)
        if response.status_code != 200 or response.status_code != 201:
            return response.json()

    def update_asset(self, asset_account):
        ac = {
            'asset_id': asset_account[0],
            'username': asset_account[3],
            'name': asset_account[3],
            'privileged': True if (asset_account[3] == 'root' or asset_account[3] == 'administrator') else False,
            'secret': asset_account[4]
        }
        rs = self.create_account_passwd(ac)
        if rs:
            print(f'新建资产账号失败：{asset_account}，接口返回：{rs}')

    def run(self):
        assets_accounts = self.get_data_from_excel()
        print(f'读取迁移数据记录个数：{len(assets_accounts)}')
        from concurrent.futures import ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=10) as executor:
            executor.map(self.update_asset, assets_accounts)


if __name__ == '__main__':
    try:
        t = Tool()
        t.run()
    except KeyboardInterrupt:
        print('用户终止程序')
